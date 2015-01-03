'''
Created on Feb 4, 2014

@author: ASALLAB
'''
from DatasetBuilder.DatasetBuilder import DatasetBuilder
import csv
import pickle
from openpyxl.workbook import Workbook
xlsxManualLabelsFileName = "C:\\Non_valeo\\Guru_Kalam\\Code\\Kalam\\DatasetBuilder\\Output\\ManualLabels_tasi_1.xlsx"
datasetSerializationFile = "C:\\Non_valeo\\Guru_Kalam\\Code\\Kalam\\DatasetBuilder\\Output\\dataset_tasi.bin"


# Reaad dataset
serializatoinDatasetFile = open(datasetSerializationFile, 'rb')
dataSet = pickle.load(serializatoinDatasetFile)
serializatoinDatasetFile.close()


# Write xlsx
import openpyxl
from openpyxl import load_workbook  

wb = Workbook()

sheet_ranges = wb.create_sheet(0, "ManualLabels_1")
row = 1
sheet_ranges.cell('A1').value = 'ID'
sheet_ranges.cell('B1').value = 'Tweet Text'
sheet_ranges.cell('C1').value = 'Relevance Label'
sheet_ranges.cell('C1').value = 'Sentiment Label'
row += 1
for item in dataSet:
    sheet_ranges.cell('A' + str(row)).value = item['id']
    sheet_ranges.cell('B' + str(row)).value = item['text']
    sheet_ranges.cell('C' + str(row)).value = item['label']
    row += 1
wb.save(xlsxManualLabelsFileName)