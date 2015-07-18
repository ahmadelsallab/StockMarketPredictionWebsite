

import re
import openpyxl
from openpyxl import load_workbook 

# Load dataset from xlsx file
def GetDatasetFromXLSXFile(fileName):
     
    dataSet = []
    try:
        # Open xlsx for reading
        wb = load_workbook(filename = fileName)
        sheet_ranges = wb.get_sheet_by_name(name = 'Sheet1')                  
        row_count = sheet_ranges.get_highest_row()
        for row in range(row_count):
            print('Reading line ' + str(row+1))
            dataSet.append({'twitter_id': sheet_ranges.cell('A' + str(row + 1)).value, 'text' : sheet_ranges.cell('B' + str(row + 1)).value})
    except openpyxl.shared.exc.InvalidFileException:
        print('File not found' + fileName)
    except Exception as e :
        print(e)
    finally:
        return dataSet
##
# Save to xlsx file
def WriteDatasetToXlsxFile(dataSet, xlsxFileName):
 
    wb = openpyxl.Workbook()

    sheet_ranges = wb.create_sheet(0, "Sheet1")
    row = 1
    for item in dataSet:
        sheet_ranges.cell('A' + str(row)).value = item['twitter_id']
        sheet_ranges.cell('B' + str(row)).value = item['text']
        row += 1
    wb.save(xlsxFileName)
    
print("Handling Dublicates")
# Read tweets from xls
#tweetes_to_render_temp = Opinion.objects.filter(stock=line, labeled = False, similarId='').values().order_by('-id')
#tweetes_to_render = sorted(tweetes_to_render_temp, key=lambda x: time.strptime(x['created_at'],'%a %b %d %X %z %Y'))
tweetes_to_render = GetDatasetFromXLSXFile("10000.xlsx")
tweets_dict = {}
tweets_dict[''] = ''
i = 1
x = 0
uniq_tweets = []
duplicate = 0
while x < len(tweetes_to_render):
    tweet_render=tweetes_to_render[x];
    tweet_render_text=tweet_render.get('text').strip()
    try:
        urls=re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', tweet_render_text);
        for i in range(0,len(urls)):
            rep='r\''+urls[i]+'\''
            tweet_render_text=re.sub(r""+urls[i]+"", '', tweet_render_text, flags=re.MULTILINE)
    except:
        pass

    tweet_render_text=re.sub(r"RT @\w*\w: ", '', tweet_render_text, flags=re.MULTILINE)
    tweet_render_text=tweet_render_text.replace("…","")
    #tweet_render_text=re.sub(r'\.\.\.', '', tweet_render_text, flags=re.MULTILINE)
    tweet_render_text=tweet_render_text[0:min(110,len(tweet_render_text))]

    if tweet_render_text in tweets_dict.keys():
        duplicate += 1
        print('Duplicate ' + str(duplicate) + ' found\n')
        
    else:
        tweets_dict[tweet_render_text] = tweet_render.get('twitter_id')
        uniq_tweets.append({'twitter_id': tweet_render['twitter_id'], 'text': tweet_render_text})
    x=x+1
WriteDatasetToXlsxFile(uniq_tweets, "10000_uniq.xlsx")
print('Final size = ' + str(len(uniq_tweets)))