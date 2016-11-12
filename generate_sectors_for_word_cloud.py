#import math

'''
x = 3
for i in range(10):
    print(str(i)+"\n")
    if(i == x):
        break
     
       '''
'''
import openpyxl
from openpyxl import load_workbook  

wb = load_workbook(filename = "C:\\Non_valeo\\Guru_Kalam\\Code\\Kalam\\DatasetBuilder\\Output\\ManualLabels_4.xlsx")

sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_4')
sheet_ranges.cell('C1').value = 'xxx'
wb.save("C:\\Non_valeo\\Guru_Kalam\\Code\\Kalam\\DatasetBuilder\\Output\\ManualLabels_4.xlsx")
'''
        
'''
size = 10
start = 0
dataSet = [0,1, 2,3, 4,5, 6,7]
while(start < dataSet.__len__()):
    remaining = dataSet.__len__() - start
    increment = min(size, remaining)
    end = start + increment
    subDataset = dataSet[start : end]
    print("XXX")
    for d in subDataset:
        print(d)
    start += increment
'''

'''
d = {'a':1, 'b':3, 'x': 4, 'd':5}
d1 = d[1:2]
for key, val in d1:# not working
    print("key: " + key + "val: " + val)
'''
'''
d = {}
d['irrel'] = 1
d['rel'] = 2
print(d['rel'])
'''

'''
reqNumRel = 2
reqNumIrrel = 2

numRel = 0
numIrrel = 0

rel = {}
irrel = {}

rel['word_1'] = 50
rel['word_2'] = 5
rel['word_3'] = 5


irrel['word_1'] = 100
irrel['word_5'] = 5
irrel['word_3'] = 5
irrel['word_4'] = 5
irrel['word_7'] = 5



finalModel = {}
for relKey, relFreq in rel.items():
    if not relKey in irrel:
        if(numRel < reqNumRel):
            finalModel[relKey] = relFreq
            numRel += 1
        
for irrelKey, irrelFreq in irrel.items():
    if not irrelKey in rel:
        if(numIrrel < reqNumIrrel):
            finalModel[irrelKey] = irrelFreq
            numIrrel += 1        
print("aa")


print(max(irrel.values()))
print(sum(irrel.values()))

a = [1, 2, 3]
x = 1
if(x in a):
    print("found")

#irrel['word_7']['w'] = 5

print(abs(-5))

'''
'''
for i in range(0,10):
    print(i)
    
    '''
    
'''
from sklearn.ensemble import AdaBoostClassifier
for i in range(0,10):
    print(i)
    
    '''
    
'''    
#print(__doc__)
import matplotlib
#matplotlib.use('Agg')
#matplotlib.rcParams['backend'] = 'wxagg'
#matplotlib.rcParams['backend'] = 'TkAgg'
matplotlib.rcParams['backend'] = 'Agg' #= matplotlib.use('Agg')
#The above fix routes the figure to anothr back end other than Tk since it has a problem with 64 bit Python 3.3
import pylab as pl
import numpy as np

from sklearn.ensemble import AdaBoostClassifier
from sklearn.tree import DecisionTreeClassifier
#from nltk import DecisionTreeClassifier
from sklearn.datasets import make_gaussian_quantiles


# Construct dataset
X1, y1 = make_gaussian_quantiles(cov=2.,
                                 n_samples=200, n_features=2,
                                 n_classes=2, random_state=1)
X2, y2 = make_gaussian_quantiles(mean=(3, 3), cov=1.5,
                                 n_samples=300, n_features=2,
                                 n_classes=2, random_state=1)
X = np.concatenate((X1, X2))
y = np.concatenate((y1, - y2 + 1))

# Create and fit an AdaBoosted decision tree

bdt = AdaBoostClassifier(DecisionTreeClassifier(max_depth=1),
                         algorithm="SAMME",
                         n_estimators=200)
'''                         

'''
Not working
bdt = AdaBoostClassifier(DecisionTreeClassifier(1),
                         algorithm="SAMME",
                         n_estimators=200)
'''
'''
bdt.fit(X, y)

plot_colors = "br"
plot_step = 0.02
class_names = "AB"

pl.figure(figsize=(10, 5))

# Plot the decision boundaries
pl.subplot(121)
x_min, x_max = X[:, 0].min() - 1, X[:, 0].max() + 1
y_min, y_max = X[:, 1].min() - 1, X[:, 1].max() + 1
xx, yy = np.meshgrid(np.arange(x_min, x_max, plot_step),
                     np.arange(y_min, y_max, plot_step))

Z = bdt.predict(np.c_[xx.ravel(), yy.ravel()])
Z = Z.reshape(xx.shape)
cs = pl.contourf(xx, yy, Z, cmap=pl.cm.Paired)
pl.axis("tight")

# Plot the training points
for i, n, c in zip(range(2), class_names, plot_colors):
    idx = np.where(y == i)
    pl.scatter(X[idx, 0], X[idx, 1],
               c=c, cmap=pl.cm.Paired,
               label="Class %s" % n)
pl.xlim(x_min, x_max)
pl.ylim(y_min, y_max)
pl.legend(loc='upper right')
pl.xlabel("Decision Boundary")

# Plot the two-class decision scores
twoclass_output = bdt.decision_function(X)
plot_range = (twoclass_output.min(), twoclass_output.max())
pl.subplot(122)
for i, n, c in zip(range(2), class_names, plot_colors):
    pl.hist(twoclass_output[y == i],
            bins=10,
            range=plot_range,
            facecolor=c,
            label='Class %s' % n,
            alpha=.5)
x1, x2, y1, y2 = pl.axis()
pl.axis((x1, x2, y1, y2 * 1.2))
pl.legend(loc='upper right')
pl.ylabel('Samples')
pl.xlabel('Decision Scores')

pl.subplots_adjust(wspace=0.25)
pl.show()
pl.savefig("plt.png")
'''



#-----------------------------------------------------------------------
# twitter-stream:
#  - ultra-real-time stream of twitter's public timeline
#    prints live results containing a URL link and the #OWS tag
#-----------------------------------------------------------------------
'''
from twitter import *

# these tokens are necessary for user authentication

# create twitter API object
auth = OAuth("1846277677-Sw5vuDcBqQdXOKZ5Pw1zLlDj8gFOfcD1L3VXx8F", "wfeTXOlxNJLXBjKBJqR6WvcD9duIjy8CrvJcsgxoqc",
                     "xNRGvHoz9L4xKGP28m7qbg", "oFv4dhBekboNg7pKa2BS0zztHqusr91SIdmKErDaycI")
stream = TwitterStream(auth = auth, secure = True)

# iterate over tweets matching this filter text
# IMPORTANT! this is not quite the same as a standard twitter search
#  - see https://dev.twitter.com/docs/streaming-api
#tweet_iter = stream.statuses.filter(track = "social")
tweet_iter = stream.statuses.sample()

for tweet in tweet_iter:
    # check whether this is a valid tweet
    if tweet.get('text'):
        # yes it is! print out the contents, and any URLs found inside
        print("(%s) @%s %s" % (tweet["created_at"], tweet["user"]["screen_name"], tweet["text"]))
        for url in tweet["entities"]["urls"]:
            print(" - found URL: %s" % url["expanded_url"])
'''

'''
import re
m = re.findall('Hello', 'Hello World')
print(m[0])

'''
#from datetime import datetime
#import time
import datetime

#print(str(datetime.time()))

import pytz


#datetime.tzinfo = pytz.timezone('Asia/Riyadh')

'''
print(' '.join(pytz.country_timezones['sa']))
riyadh_time = pytz.timezone('Asia/Riyadh')
now_time_local = datetime(datetime.now(), tzinfo=pytz.utc) 
now_time_riyadh = riyadh_time.localize(now_time_local, is_dst=True)
fmt = '%Y-%m-%d %H:%M:%S %Z%z'
print(now_time_riyadh.strftime(fmt))
#for key, val in pytz.country_names:
#for key, val in pytz.country_timezones:
#    print(key + "  ", val) 
print(str(datetime.now().time()))
print(str(now_time_riyadh.time()))
'''
#open_time = datetime.time(11,0,0, tzinfo=pytz.timezone('Asia/Riyadh'));
#open_time = datetime.timetz(11,0,0)

'''
print(datetime.datetime.now(tz=pytz.timezone('Asia/Riyadh')))
print(datetime.datetime.now(tz=pytz.timezone('Asia/Riyadh')).time())

now_dt = datetime.datetime.utcnow()
date_dt = now_dt.date()
t = datetime.time(11,0,0)
dt = datetime.datetime.combine(date_dt, t)
dt.replace(tzinfo = pytz.timezone(str('Asia/Riyadh')))
print(dt.time())

# t = datetime.time(datetime.datetime.timetz())

print(t)
'''
import urllib.request
from bs4 import BeautifulSoup
f1 = urllib.request.urlopen('http://www.marketstoday.net/markets/%D8%A7%D9%84%D8%B3%D8%B9%D9%88%D8%AF%D9%8A%D8%A9/Companies/1/ar/')


soup = BeautifulSoup(f1.read())  
f = open('stocks_sectors.txt', 'w', encoding = 'utf-8')
sectors_id = []
sectors = []
for a in soup.findAll('span', { "class" : "compsect" }):
    #print('Sector' + a.text.strip())
    #f.write('Sector: ' + a.text.strip() + '\n')
    sectors.append(a.text.strip())
    
for a in soup.findAll('div', { "class" : "compexpand" }):
    #f.write(a.attrs['name'])
    sectors_id.append(a.attrs['id'])
i = 0
'''
for sector in sectors:
    f.write('Sector: ' + sector + '\n')
    sector_html = soup.findAll('div', { "id" : sectors_id[i] })
    for company in sector_html[0].findAll('a', { "class" : "jTip" }):
        f.write('Company: ' + company.text.strip() + '\n')
    i = i + 1
'''
f.write('stocks_sectors = {')
#"تاسي":["industry"]}
for sector in sectors:
    sector_html = soup.findAll('div', { "id" : sectors_id[i] })
    for company in sector_html[0].findAll('a', { "class" : "jTip" }):
        f.write('\'' + company.text.strip() + '\':\'' + sector + '\', \n')
    i = i + 1 
f.write('}')
f.close()
#http://www.w3schools.com/html/html_colors.asp
sectors_colors = ["#000000", "#000033", "#000066", "#000099", "#0000CC", "#0000FF",
"#003300", "#003333", "#003366", "#003399", "#0033CC", "#0033FF",
"#006600", "#006633", "#006666", "#006699", "#0066CC", "#0066FF",
"#009900", "#009933", "#009966", "#009999", "#0099CC", "#0099FF",
"#00CC00", "#00CC33", "#00CC66", "#00CC99", "#00CCCC", "#00CCFF",
"#00FF00", "#00FF33", "#00FF66", "#00FF99", "#00FFCC", "#00FFFF",
"#330000", "#330033", "#330066", "#330099", "#3300CC", "#3300FF",
"#333300", "#333333", "#333366", "#333399", "#3333CC", "#3333FF",
"#336600", "#336633", "#336666", "#336699", "#3366CC", "#3366FF",
"#339900", "#339933", "#339966", "#339999", "#3399CC", "#3399FF",
"#33CC00", "#33CC33", "#33CC66", "#33CC99", "#33CCCC", "#33CCFF",
"#33FF00", "#33FF33", "#33FF66", "#33FF99", "#33FFCC", "#33FFFF",
"#99CC00", "#99CC33", "#99CC66", "#99CC99", "#99CCCC", "#99CCFF",
"#99FF00", "#99FF33", "#99FF66", "#99FF99", "#99FFCC", "#99FFFF",
"#CC0000", "#CC0033", "#CC0066", "#CC0099", "#CC00CC", "#CC00FF",
"#CC3300", "#CC3333", "#CC3366", "#CC3399", "#CC33CC", "#CC33FF",
"#CC6600", "#CC6633", "#CC6666", "#CC6699", "#CC66CC", "#CC66FF",
"#CC9900", "#CC9933", "#CC9966", "#CC9999", "#CC99CC", "#CC99FF",
"#CCCC00", "#CCCC33", "#CCCC66", "#CCCC99", "#CCCCCC", "#CCCCFF",
"#CCFF00", "#CCFF33", "#CCFF66", "#CCFF99", "#CCFFCC", "#CCFFFF",
"#FF0000", "#FF0033", "#FF0066", "#FF0099", "#FF00CC", "#FF00FF",
"#FF3300", "#FF3333", "#FF3366", "#FF3399", "#FF33CC", "#FF33FF",
"#FF6600", "#FF6633", "#FF6666", "#FF6699", "#FF66CC", "#FF66FF",
"#FF9900", "#FF9933", "#FF9966", "#FF9999", "#FF99CC", "#FF99FF",
"#FFCC00", "#FFCC33", "#FFCC66", "#FFCC99", "#FFCCCC", "#FFCCFF",
"#FFFF00", "#FFFF33", "#FFFF66", "#FFFF99", "#FFFFCC", "#FFFFFF",
"#FFFF00", "#FFFF33", "#FFFF66", "#FFFF99"]

f = open('jqcloud.css', 'w', encoding = 'utf-8')

j = 0
print(len(sectors_colors))
for sector in sectors:
    for i in range(10):
        f.write('.' + sector + '.w' + str(10 - i) + '{ color: #' + sectors_colors[j] + '; }\n')
        j = j + 1

f.close()