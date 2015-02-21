'''
Created on Oct 23, 2013

@author: ASALLAB
'''
import random
import csv
import pickle
from xml.dom import minidom
import openpyxl
from openpyxl import load_workbook
DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST = False
DEBUG_LIMIT_IRRELEVANT_TRAIN_ONLY = False
class DatasetBuilder(object):
    '''
    classdocs
    '''
    
    # The dataset after necessary pre-processing
    dataSet = []

    # The raw data
    rawData = []

    # Colomn index of the relevance label
    RELEVANCE_LABEL_COL_IDX = 2
        
    def __init__(self, configFileName, rawDataSetsList, datasetSerializationFile):
        '''
        Constructor
        '''
        # The raw data
        self.rawDataSetsList = rawDataSetsList
        
        # The processed dataset
        self.dataSet = []
                
        # Save the dataset serialization file
        self.datasetSerializationFile = datasetSerializationFile
        
        # Parse the configurations file
        self.ParseConfigFile(configFileName)
        if(DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST):
            self.irrelevantNum = 0
            self.limitIrrelevant = 5000
            #/DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
    def BuildDataset(self):
        
        # For each dataset with label in the rawDataSetList, get a random dataSetSize list
        # And join them in one list
        joinedList = []
        if(self.balanceUserShare == "false"):
            for dataset in self.rawDataSetsList:  
                joinedList.extend(self.RandomizeAndLabel(dataset['label'], dataset['rawData'], self.datasetSize))          
        else:
            for dataset in self.rawDataSetsList:
                joinedList.extend(self.BalanceUsersShare(dataset['label'], dataset['rawData'], self.datasetSize))
        
        # Choose randomly dataSetSize examples of the joinedList
        finalList = self.RandomizeList(joinedList, self.datasetSize)
        self.dataSet = finalList
            
    def BalanceUsersShare(self, label, origDataset, desiredSize):   
        
        # Convert into users DB:
        #-----------------------
        # The users list
        users = {}
        
        balancedDataset = []
           
        # The desired dataset size is the min. of the rawData and the configuration parameter  
        size = min(desiredSize,origDataset.__len__())    
        
        usersSize = 0
        # Loop on all data
        for rawDataEntry in origDataset:
            
            # Insert the user if not already existing
            #users[rawDataEntry['user']['id']] = rawDataEntry['user']['id']
            
            
            try:
                # Insert a new tweet text
                users[rawDataEntry['user']['id']]['tweets'].append(rawDataEntry)
                
                # Increment the remaining tweets
                users[rawDataEntry['user']['id']]['remainingTweets'] += 1 
            except:
                # Initialize the user entry as empty dict
                users[rawDataEntry['user']['id']] = {}
                usersSize += 1
                
                # Initialize the list of user tweets and insert the first tweet
                users[rawDataEntry['user']['id']]['tweets'] = []
                users[rawDataEntry['user']['id']]['tweets'].append(rawDataEntry)
                
                # Initialize the remaining tweets
                users[rawDataEntry['user']['id']]['remainingTweets'] = 1
                
                # Initialize the tweets share per user
                users[rawDataEntry['user']['id']]['tweetsShare'] = 0
        
        # Balance users share and build dataset
        #--------------------------------------
        
        # Initialize the unbalanced data size
        remainingDataSize = size
        
        # Initialize the remaining users not balanced
        remNumUsers = users.__len__()
        
        # Loop until no more data still unbalanced
        # The second check remNumUsers >= 1 is added because in some cases, the user share is fraction:
        # userFairShare = max((remainingDataSize / remNumUsers), 1)
        # Like numUsers = 978 and remainingDataSize = 1000, then each user is required to provide 1.02 tweet, so 1
        # Then we'll have 978 tweets from 978 user and remaining 22 in remainingDataSize. Hence the loop will never exit
        while remainingDataSize > 0 and remNumUsers >= 1:
            # Loop on all users and balance their shares
            for userId, user in users.items():
               
                # Stop immediately when the remaining size to be balanced is 0
                if not (remainingDataSize > 0):
                    break
                # Increment the user share based on the minimum increment, either its full data or equally divide the shares among remaining users
                # The user fair share is to divide equally among users. If the share is less than 1 tweet per user, take 1 tweet from each.
                # This happens when the remainingDataSize is less than remNumUsers
                # A user is processed only if there are remaining tweets to be included in his list
                if user['remainingTweets'] > 0:
                    
                    userFairShare = max((remainingDataSize / remNumUsers), 1)
                        
                    increment = int(min(userFairShare, user['remainingTweets']))
                    user['tweetsShare'] += increment
                    
                    # Discount the remaining unbalanced data
                    remainingDataSize -= increment                                        
                    
                    # Discount the remaining unbalanced tweets from the user
                    user['remainingTweets'] -= increment
                
                    # Remove the user from the remaining users if all its tweets are balanced
                    if(user['remainingTweets'] <= 0):
                        # Remove the user
                        remNumUsers -= 1
                        
                        # Add to the dataset random selection of the user tweets until its final share is consumed
                        
                        # Create random indices
                        try:
                            randIndices = random.sample(range(user['tweets'].__len__()), user['tweetsShare'])
                        except:
                            print(1)
                        
                        # Insert in the final dataset N=user.tweetsShare random tweets from the user.tweets
                        for index in randIndices:
                            balancedDataset.append(user['tweets'][index])
        
        # Return just list of 'text' and 'label'
        unLabeledFinalList = []
        for item in balancedDataset:
            unLabeledFinalList.append(item)
            
        # Label the dataset
        labeledFinalList = self.LabelList(unLabeledFinalList, label)
        
        return labeledFinalList
    # Utility to randomize a list
    def RandomizeList(self, origList, size):
        
        randomList = []
        # Limit the size
        randomListSize = min(origList.__len__(), size)
        
        # Build the dataset randomly just from the original raw
        # Create random indices
        randIndices = random.sample(range(origList.__len__()), randomListSize)
        
        # Insert in the final dataset N=self.datasetSize random tweets from the rawData
        for index in randIndices:
            randomList.append(origList[index])
        
        return randomList       
    # Utility to randomize a list
    def LabelList(self, origList, label):       
        # Add a new item in the list of dict items called 'label'
        for item in origList:
            item['label'] = label
        return origList
    
    def RandomizeAndLabel(self, label, origDataset, desiredSize):
        
        # The desired dataset size is the min. of the rawData and the configuration parameter  
        size = min(desiredSize, origDataset.__len__())    
                
        # First randomize the list and get the required size
        randomList = self.RandomizeList(origDataset, size)
        
        unLabeledFinalList = []
        
        # Build list of text fields only
        for item in randomList:
            unLabeledFinalList.append(item)  
        
        # Label the dataset       
        labeledFinalList = self.LabelList(unLabeledFinalList, label)
        
        return labeledFinalList
        
    # Parse the configurations file
    def ParseConfigFile(self, configDocName):
        # Get the name of configuration file from the cmd line argument
        xmldoc = minidom.parse(configDocName)    
        
        # Get the user share balance option
        self.balanceUserShare = xmldoc.getElementsByTagName('BalanceUserShare')[0].attributes['balanceUserShare'].value

        # Get the desired dataset size
        self.datasetSize = int(xmldoc.getElementsByTagName('DatasetSize')[0].attributes['datasetSize'].value)
        
        # Get the testSetShare
        self.testSetShare = float(xmldoc.getElementsByTagName('TestSetShare')[0].attributes['testSetShare'].value)
        
        # Get the NumberTweetsPerCsvFile
        numberTweetsPerCsvFile = xmldoc.getElementsByTagName('NumberTweetsPerCsvFile')[0].attributes['numberTweetsPerCsvFile'].value
        if(numberTweetsPerCsvFile == "All"):
            # All dataset shall be dumped in one file
            self.numberTweetsPerCsvFile = self.datasetSize
        else:
            self.numberTweetsPerCsvFile = int(numberTweetsPerCsvFile)
        
            

    # Dump text to be manually labeled to csv
    def DumpDatasetToCSV(self, csvManualLabelsFileName):
        
        csvFileIndex = 1
        numItems = 0
        
        # Open the first file
        f = open((csvManualLabelsFileName + "_" + str(csvFileIndex) + ".csv"), 'w', encoding='utf-8', newline='')
        
        # Get writer handler, ',' delimited
        w = csv.writer(f, delimiter=',')
         
        # Fill in the data rows
        # Start the heading row
        data = []
        data.append(['ID', 'Tweet text', 'Relevance label', 'Sentiment label'])

        # Fill in each row with the dataset initial labels
        for item in self.dataSet:
            data.append([item['id'], item['text'], item['label'], ' '])
            numItems += 1
            if(numItems >= self.numberTweetsPerCsvFile):
                # Dump to the csv file
                w.writerows(data)
                
                # Close the old file
                f.close()
                
                # Start new file
                numItems = 0
                csvFileIndex += 1
                
                # Open the next file
                f = open((csvManualLabelsFileName + "_" + str(csvFileIndex) + ".csv"), 'w', encoding='UTF-8', newline='')
                
                # Get writer handler, ',' delimited
                w = csv.writer(f, delimiter=',')
                 
                # Fill in the data rows
                # Start the heading row
                data = []
                data.append(['ID', 'Tweet text', 'Relevance label', 'Sentiment label'])
              
        # If the file is still open, then dump and close it  
        if(not f.closed):
            # Dump to the csv file
            w.writerows(data)
            
            # Close the file
            f.close()
                  
                
    # Dump text to be manually labeled to xlsx
    def DumpDatasetToXLSX(self, xlsxManualLabelsFileName):
        
        start = 0
        xlsxFileIndex = 1
        while(start < self.dataSet.__len__()):
            remaining = self.dataSet.__len__() - start
            increment = min(self.numberTweetsPerCsvFile, remaining)
            end = start + increment
            subDataset = self.dataSet[start : end]
            xlsxFileName = (xlsxManualLabelsFileName + "_" + str(xlsxFileIndex) + ".xlsx")
            self.WriteDatasetToXlsxFile(subDataset, xlsxFileName, xlsxFileIndex)
            start += increment
            xlsxFileIndex += 1
        
            
    # Save to xlsx file
    def WriteDatasetToXlsxFile(self, dataSet, xlsxFileName, index):
        import openpyxl
        from openpyxl import load_workbook  
        wb = openpyxl.Workbook()

        sheet_ranges = wb.create_sheet(0, "ManualLabels_" + str(index))
        row = 1
        sheet_ranges.cell('A1').value = 'ID'
        sheet_ranges.cell('B1').value = 'Tweet text'
        sheet_ranges.cell('C1').value = 'Relevance label'
        sheet_ranges.cell('D1').value = 'Sentiment label'
        row += 1
        for item in dataSet:
            sheet_ranges.cell('A' + str(row)).value = item['id']
            sheet_ranges.cell('B' + str(row)).value = item['text']
            sheet_ranges.cell('C' + str(row)).value = item['label']
            row += 1
        wb.save(xlsxFileName)            
    # Update labels from manually labeled csv file
    def UpdateManualLabelsFromCSV(self, csvManualLabelsFileName):
        
        # Open csv for reading
        f = open(csvManualLabelsFileName + ".csv", 'r', encoding='UTF-8', newline='')
        
        # Get reader handler
        r = csv.reader(f, delimiter=',')
        
        # Start at the first entry of the dataset
        # By default, the rows are ordered similar to the order of the original dataset
        index = 0
        
        # Skip the first row
        skip = True
        # Read the label from each row
        for row in r:
            if(not skip):
                # Update the label from the manual updated labels
                self.dataSet[index]['label'] = row[self.RELEVANCE_LABEL_COL_IDX]
                self.dataSet[index]['text'] = row[1]
                index += 1
            else:
                skip = False
        # Close the file   
        f.close()
        
    # Load dataset from xlsx file
    def LoadDatasetFromXLSXFile(self, fileName):
         
        try:
            # Open xlsx for reading
            wb = load_workbook(filename = fileName + ".xlsx")
            sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_1')
                       
            # Read the label from each row
            for row in range(self.numberTweetsPerCsvFile + 2): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped
                if(row >= 2):
                    # Update the label from the manual updated labels
                    #NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if not DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                    #/NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    
                    #DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        if(sheet_ranges.cell('C' + str(row)).value == 'irrelevant'):
                            self.irrelevantNum += 1
                            if(self.irrelevantNum <= self.limitIrrelevant):
                                # Update the label from the manual updated labels
                                self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                        else:
                                self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                            
                    #/DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST        
        except openpyxl.shared.exc.InvalidFileException:
            print('File not found' + fileName + ".xlsx")
            
    # Load dataset from xlsx file
    def GetDatasetFromXLSXFile(self, fileName):
         
        dataSet = []
        try:
            # Open xlsx for reading
            wb = load_workbook(filename = fileName + ".xlsx")
            sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_1')
                      
            row_count = sheet_ranges.get_highest_row()
            # Read the label from each row            
            #for row in range(self.numberTweetsPerCsvFile + 2): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped
            for row in range(row_count): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped                if(row >= 2):
                if(row >= 2):
                    # Update the label from the manual updated labels
                    #NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if not DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                    #/NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    
                    #DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        if(sheet_ranges.cell('C' + str(row)).value == 'irrelevant'):
                            self.irrelevantNum += 1
                            if(self.irrelevantNum <= self.limitIrrelevant):
                                # Update the label from the manual updated labels
                                dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                        else:
                                dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                            
                    #/DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
        except openpyxl.shared.exc.InvalidFileException:
            print('File not found' + fileName + ".xlsx")
        finally:
            return dataSet
    # Load dataset from xlsx file
    def GetDatasetFromXLSXFileBipolarSentiment(self, fileName):
         
        dataSet = []
        try:
            # Open xlsx for reading
            wb = load_workbook(filename = fileName + ".xlsx")
            sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_1')
                      
            row_count = sheet_ranges.get_highest_row()
            # Read the label from each row            
            #for row in range(self.numberTweetsPerCsvFile + 2): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped
            for row in range(row_count): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped                if(row >= 2):
                if(row >= 2):
                    # Update the label from the manual updated labels
                    #NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if not DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        dataSet.append({'label' : sheet_ranges.cell('E' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                    #/NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    
                    #DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        if(sheet_ranges.cell('C' + str(row)).value == 'irrelevant'):
                            self.irrelevantNum += 1
                            if(self.irrelevantNum <= self.limitIrrelevant):
                                # Update the label from the manual updated labels
                                dataSet.append({'label' : sheet_ranges.cell('E' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                        else:
                                dataSet.append({'label' : sheet_ranges.cell('E' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                            
                    #/DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
        except openpyxl.shared.exc.InvalidFileException:
            print('File not found' + fileName + ".xlsx")
        finally:
            return dataSet

    # Update labels from manually labeled xlsx file
    def UpdateManualLabelsFromXLSXFile(self, fileName, index):
         
        try:
            # Open xlsx for reading
            wb = load_workbook(filename = fileName + ".xlsx")
            sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_' + str(index))
            #sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_1')
           
            # Read the label from each row
            for row in range(self.numberTweetsPerCsvFile + 2): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped
                if(row >= 2):
                    # Update the label from the manual updated labels
                    #NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if not DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                    #/NOT DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    
                    #DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST
                    if DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST:
                        if(sheet_ranges.cell('C' + str(row)).value == 'irrelevant'):
                            self.irrelevantNum += 1
                            if(self.irrelevantNum <= self.limitIrrelevant):
                                # Update the label from the manual updated labels
                                self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                        else:
                                self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                            
                    #/DEBUG_LIMIT_IRRELEVANT_TRAIN_AND_TEST        
        except openpyxl.shared.exc.InvalidFileException:
            print('File not found' + fileName + ".xlsx")
    # Update labels from manually labeled xlsx file
    def UpdateManualLabelsFromXLSXFileFilterByQuery(self, fileName, index, queryArray):
        import re
        try:
            # Open xlsx for reading
            wb = load_workbook(filename = fileName + ".xlsx")
            sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels' + str(index))
            #sheet_ranges = wb.get_sheet_by_name(name = 'ManualLabels_1')
           
            # Read the label from each row
            for row in range(self.numberTweetsPerCsvFile + 2): # +2 since the xlsx rows start at 1 and also the first row is the header and should be skipped
                if(row >= 2):
                    # Update the label from the manual updated labels
                    for query in queryArray:
                        try:
                            if(re.findall('\\b' + query + '\\b', sheet_ranges.cell('B' + str(row)).value).__len__() != 0):
                                self.dataSet.append({'label' : sheet_ranges.cell('C' + str(row)).value, 'text' : sheet_ranges.cell('B' + str(row)).value})
                                '''
                                if(sheet_ranges.cell('C' + str(row)).value == 'irrelevant'):
                                    print(sheet_ranges.cell('B' + str(row)).value)
                                else:
                                    print(sheet_ranges.cell('B' + str(row)).value)
                                '''
                                break # to get out of the loop, you insert only in case the one query at least exist in the text
                                
                        except TypeError:
                            print('Error ' + str(sheet_ranges.cell('B' + str(row)).value))
                    
        except openpyxl.shared.exc.InvalidFileException:
            print('File not found' + fileName + ".xlsx")
            
    # Update labels from manually labeled csv file
    def UpdateManualLabelsFromCSVFile(self, fileName):
        
        # Open csv for reading
        f = open(fileName + ".csv", 'r', encoding='UTF-8', newline='')
        #f = open(fileName + ".csv", 'r', newline='')
        #f = open(fileName + ".xlsx", 'r', encoding='UTF-8', newline='')
        
        # Get reader handler
        r = csv.reader(f, delimiter=',')
        
        
        # Skip the first row
        skip = True
        # Read the label from each row
        for row in r:
            if(not skip):
                # Update the label from the manual updated labels
                self.dataSet.append({'label' : row[self.RELEVANCE_LABEL_COL_IDX], 'text' : row[1]})
            else:
                skip = False
        # Close the file   
        f.close()        

    # To save to serialzation file
    def SaveDataset(self):
        # You must close and open to append to the binary file
        # Open the serialization file
        self.serializationFile = open(self.datasetSerializationFile, 'wb')
        
        #pickle.dump(self.dataSet, self.datasetSerializationFile)
        pickle.dump(self.dataSet, self.serializationFile)
        
        # Open the serialization file
        self.serializationFile.close()

        
    # To load saved dataset
    def LoadDataset(self):
        # Load the dataset
        serializatoinDatasetFile = open(self.datasetSerializationFile, 'rb')
        self.dataSet = pickle.load(serializatoinDatasetFile)
        serializatoinDatasetFile.close()
        
    # Internal utility to split a dataset list randomly into two sets: train and test, according to the configured desired test share
    # testsetShare must be < 1        
    def SplitTrainTest(self):
        self.trainSet, self.testSet =  self.SplitTrainTestInternal(self.dataSet, self.testSetShare)
                    
    # Internal utility to split a dataset list randomly into two sets: train and test, according to the desired test share
    # testsetShare must be < 1
    def SplitTrainTestInternal(self, dataSet, testSetShare):
        
        # Randomize the dataSet
        randomDataSet = self.RandomizeList(dataSet, dataSet.__len__())
        
        # Calculate the test set size
        testSetSize = int(testSetShare * dataSet.__len__())
        
        # The trainSet starts from the begining until before the end by the test set size 
        trainSet = randomDataSet[0:dataSet.__len__() - testSetSize]
        if(DEBUG_LIMIT_IRRELEVANT_TRAIN_ONLY):
            irrelevantSize = 60000
            trainSet_ = trainSet
            trainSet = []
            for example in trainSet_:
                if example['label'] == 'irrelevant':
                    irrelevantSize -= 1
                    if irrelevantSize >= 0:
                        trainSet.append(example)
                else:
                    trainSet.append(example)
            # The testSet is the last number of testSetSize examples
            testSet = randomDataSet[trainSet_.__len__():]
        else:                           
            # The testSet is the last number of testSetSize examples
            testSet = randomDataSet[trainSet.__len__():]
        
        return trainSet, testSet
                
    # To save to serialzation file
    def SaveTrainTestDataset(self, fileName):
        # You must close and open to append to the binary file
        # Open the serialization file
        self.serializationFile = open(fileName, 'wb')

        pickle.dump(self.trainSet, self.serializationFile)
        pickle.dump(self.testSet, self.serializationFile)
        
        # Open the serialization file
        self.serializationFile.close()

        
    # To load saved dataset
    def LoadTrainTestDataset(self, fileName):
        # Load the dataset
        serializatoinDatasetFile = open(fileName, 'rb')
        self.trainSet = pickle.load(serializatoinDatasetFile)
        self.testSet = pickle.load(serializatoinDatasetFile)
        serializatoinDatasetFile.close()